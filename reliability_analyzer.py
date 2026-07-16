#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Discrete Reliability Data Analyzer
==================================
경량 스택: Tkinter + Matplotlib + csv/openpyxl (Pandas/PySide6 미사용)
- 여러 Read-out 측정 파일(CSV/XLSX)을 읽어 Parameter별 변화/산포 분석
- 신뢰성명 × Lot 조합(그룹)별로 완전히 구분하여 각각 그래프 생성
- Read-out graph / Delta % graph / Box plot + 편집(Undo/Redo) + PDF Report
"""

import csv
import math
import os
import re
import sys
import threading
import traceback

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, colorchooser, simpledialog

import matplotlib
matplotlib.use("TkAgg")
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.backends.backend_pdf import PdfPages
from matplotlib.ticker import MultipleLocator
import matplotlib.transforms as mtransforms

# Drag & Drop (선택적 — 미설치 시 Browse만 동작)
try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
    HAS_DND = True
except Exception:
    HAS_DND = False

# ----------------------------------------------------------------------------
# 제한사항 (그룹당)
# ----------------------------------------------------------------------------
MAX_SAMPLES = 500
MAX_PARAMS = 500
MAX_READOUTS = 20

READOUT_COLORS = [
    "#1f77b4", "#ff7f0e", "#2ca02c", "#d62728", "#9467bd",
    "#8c564b", "#e377c2", "#7f7f7f", "#bcbd22", "#17becf",
    "#aec7e8", "#ffbb78", "#98df8a", "#ff9896", "#c5b0d5",
    "#c49c94", "#f7b6d2", "#c7c7c7", "#dbdb8d", "#9edae5",
]


def col_title(col):
    """컬럼명 'ITEM@BIAS (UNIT)[ #n]' → 'ITEM [UNIT] @BIAS[ #n]' 표시 형식."""
    m = re.match(r"^(.*?)@(.*?)\s*\(([^()]*)\)(\s*#\d+)?$", col)
    if not m:
        return col
    item, bias, unit, suf = m.group(1), m.group(2), m.group(3), m.group(4) or ""
    return f"{item} [{unit}] @{bias}{suf}"


def center_window(win, w=None, h=None):
    """창을 화면 정중앙에 배치."""
    win.update_idletasks()
    ww = w or win.winfo_width() or win.winfo_reqwidth()
    wh = h or win.winfo_height() or win.winfo_reqheight()
    x = (win.winfo_screenwidth() - ww) // 2
    y = (win.winfo_screenheight() - wh) // 2
    win.geometry(f"+{x}+{y}")


# ============================================================================
# 1. 파일명 파싱: 신뢰성명 + Lot번호 + Read-out (구분자 _ - + 공백, 순서 무관)
# ============================================================================
READOUT_RE = re.compile(r"^(\d+(?:\.\d+)?)\s*(hr|hrs|h|hour|hours|cyc|cycle|cycles|cy)$", re.I)
LOT_RE = re.compile(r"^lot\s*([A-Za-z0-9]+)$", re.I)


def parse_filename(path):
    """파일명에서 (reliability, lot, readout_label, readout_value) 추출.
    실패 시 ValueError."""
    base = os.path.splitext(os.path.basename(path))[0]
    tokens = [t for t in re.split(r"[_\s]+", base) if t]
    readout_label = None
    readout_value = None
    lot = None
    rest = []
    is_retest = False
    for tok in tokens:
        if tok.lower() == "retest":
            is_retest = True
            continue
        m = READOUT_RE.match(tok)
        if m and readout_label is None:
            readout_value = float(m.group(1))
            unit = m.group(2).lower()
            unit = "hr" if unit.startswith("h") else "cyc"
            num = m.group(1)
            readout_label = f"{num}{unit}"
            continue
        m = LOT_RE.match(tok)
        if m and lot is None:
            lot = "LOT" + m.group(1).upper()
            continue
        rest.append(tok)
    if readout_label is None or lot is None or not rest:
        raise ValueError(
            f"파일명 인식 실패: '{os.path.basename(path)}'\n"
            "파일 이름은 신뢰성명 + Lot번호 + Read-out 형식이어야 합니다.\n"
            "예: HTRB_Lot1_0hr, HTBG+_Lot2_500cyc (구분자는 _ 또는 공백)"
        )
    reliability = "_".join(rest).upper()
    if is_retest:
        reliability = "RETEST_" + reliability
    return reliability, lot, readout_label, readout_value


# ============================================================================
# 2. 데이터 파싱 (좌표 규칙: Column6 키워드 탐색)
# ============================================================================
def _read_rows(path):
    """CSV/XLSX → 모든 셀 strip()된 문자열 2차원 리스트."""
    ext = os.path.splitext(path)[1].lower()
    rows = []
    if ext == ".csv":
        for enc in ("utf-8-sig", "cp949", "latin-1"):
            try:
                with open(path, newline="", encoding=enc) as f:
                    rows = [[(c or "").strip() for c in r] for r in csv.reader(f)]
                break
            except UnicodeDecodeError:
                rows = []
                continue
    elif ext in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        for r in ws.iter_rows(values_only=True):
            rows.append([("" if c is None else str(c)).strip() for c in r])
        wb.close()
    else:
        raise ValueError(f"지원하지 않는 확장자: {ext}")
    return rows


def _cell(row, idx):
    return row[idx].strip() if idx < len(row) else ""


def parse_data_file(path):
    """좌표 규칙에 따라 파일 파싱.
    반환: (columns, data)
      columns: [f"Item@Bias1 (Unit)"] — Unit이 빈 열 제외, 중복명은 #2 접미사
      data: {sample_no(int): {colname: float|None}}"""
    fname = os.path.basename(path)
    rows = _read_rows(path)

    item_row = bias_row = unit_row = None
    data_start = None
    for i, row in enumerate(rows):
        c6 = _cell(row, 6)
        if item_row is None and c6 == "Item":
            item_row = row  # 첫 번째 Item 행만 사용 (통계 블록의 두 번째 Item 무시)
        elif c6 == "Bias1" and bias_row is None:
            bias_row = row
        elif c6 == "Unit" and unit_row is None:
            unit_row = row
        if data_start is None and _cell(row, 0) == "Test No.":
            data_start = i + 1

    missing = []
    if item_row is None:
        missing.append("Item")
    if bias_row is None:
        missing.append("Bias1")
    if unit_row is None:
        missing.append("Unit")
    if data_start is None:
        missing.append("Test No.")
    if missing:
        raise ValueError(f"'{fname}' 에서 {', '.join(missing)} 행을 찾을 수 없습니다.")

    ncol = max(len(item_row), len(bias_row), len(unit_row))
    col_map = []
    name_count = {}
    for j in range(7, ncol):
        item = _cell(item_row, j)
        unit = _cell(unit_row, j)
        bias = _cell(bias_row, j)
        if not item or not unit:
            continue
        name = f"{item}@{bias} ({unit})"
        if name in name_count:
            name_count[name] += 1
            name = f"{name} #{name_count[name]}"
        else:
            name_count[name] = 1
        col_map.append((name, j))

    data = {}
    for row in rows[data_start:]:
        s = _cell(row, 0)
        if not s:
            continue
        try:
            sample = int(float(s))
        except ValueError:
            continue
        vals = {}
        for colname, j in col_map:
            v = _cell(row, j)
            try:
                vals[colname] = float(v)
            except ValueError:
                vals[colname] = None
        data[sample] = vals

    columns = [c for c, _ in col_map]
    return columns, data


# ============================================================================
# 3. 데이터 모델 — 그룹(신뢰성명_Lot) 단위로 완전 구분
# ============================================================================
class DataModel:
    """
    groups: 정렬된 그룹명 목록. 그룹 = f"{신뢰성명}_{LOT}"
    g[group] = {
        'readouts': 정렬된 Read-out label 목록,
        'columns' : Parameter 컬럼명 목록,
        'samples' : 정렬된 시료 번호 목록,
        'data'    : data[readout][sample][col] = float|None
    }
    편집 상태 키에는 항상 group이 포함되어 그룹 간 간섭이 없음.
    """

    def __init__(self):
        self.reliability = None
        self.groups = []   # Lot 목록 (예: ['LOT1', 'LOT2'])
        self.g = {}
        self.deleted = set()        # (group, readout, col, sample)
        self.color_over = {}        # (group, readout, col, sample) -> hex
        self.ylim = {}              # (group, col) -> (ymin, ymax)
        self.limits = {}            # (group, col) -> {'high': v|None, 'low': v|None}
        self.custom_lines = {}      # (group, col, kind) -> [ {axis,value,style,color,name} ]
        self.texts = {}             # (group, col, kind) -> [ {text,x,y,size,color} ]
        self._undo = []
        self._redo = []

    # ---- 로드 ------------------------------------------------------------
    def load(self, files):
        """files 파싱. 반환: errors(list[str]). 그룹별로 완전히 분리 적재."""
        errors = []
        # raw[lot] = list of (readout_label, readout_value, columns, data, fname)
        raw = {}
        rels = {}
        for p in files:
            fname = os.path.basename(p)
            try:
                rel, lot, rl, rv = parse_filename(p)
                cols, d = parse_data_file(p)
            except ValueError as e:
                errors.append(str(e))
                continue
            rels.setdefault(rel, []).append(fname)
            raw.setdefault((rel, lot), []).append((rl, rv, cols, d, fname))

        if not raw:
            return errors

        # Retest_X와 X는 같은 신뢰성(별도 분석 단위). base가 다르면 차단
        bases = {}
        for r, fs in rels.items():
            base = r[7:] if r.upper().startswith("RETEST_") else r
            bases.setdefault(base, []).extend(fs)
        if len(bases) > 1:
            detail = "\n".join(f"  {b}: {', '.join(fs)}" for b, fs in bases.items())
            errors.append("한 번에 하나의 신뢰성만 분석할 수 있습니다.\n"
                          f"여러 신뢰성이 섞여 있습니다:\n{detail}")
            return errors
        self.reliability = next(iter(bases))

        self.groups = []
        self.g = {}
        for rel_key, lot in sorted(raw):
            group = lot if not rel_key.upper().startswith("RETEST_") else f"{lot}(R)"
            entries = raw[(rel_key, lot)]
            # 동일 그룹 내 동일 Read-out 중복 → 오류로 안내 (덮어쓰기 방지)
            seen_ro = {}
            dup = []
            for rl, rv, cols, d, fname in entries:
                if rl in seen_ro:
                    dup.append(f"{group}의 {rl}: '{seen_ro[rl]}' 와 '{fname}'")
                else:
                    seen_ro[rl] = fname
            if dup:
                errors.append("동일 그룹에 같은 Read-out 파일이 중복되었습니다:\n"
                              + "\n".join(dup))
                continue
            if len(entries) > MAX_READOUTS:
                errors.append(f"{group}: Read-out 파일이 {MAX_READOUTS}개를 "
                              f"초과합니다 ({len(entries)}개).")
                continue

            entries.sort(key=lambda x: x[1])
            readouts = [e[0] for e in entries]
            col_union, seen = [], set()
            samp_union = set()
            data = {}
            for rl, rv, cols, d, fname in entries:
                for c in cols:
                    if c not in seen:
                        seen.add(c)
                        col_union.append(c)
                samp_union.update(d.keys())
                data[rl] = d

            if len(col_union) > MAX_PARAMS:
                errors.append(f"{group}: Parameter가 {MAX_PARAMS}개를 초과합니다 "
                              f"({len(col_union)}개).")
                continue
            if len(samp_union) > MAX_SAMPLES:
                errors.append(f"{group}: Sample이 {MAX_SAMPLES}개를 초과합니다 "
                              f"({len(samp_union)}개).")
                continue

            self.groups.append(group)
            self.g[group] = dict(readouts=readouts, columns=col_union,
                                 samples=sorted(samp_union), data=data,
                                 rel=rel_key, lot=lot)

        self.deleted.clear()
        self.color_over.clear()
        self.ylim.clear()
        self.limits.clear()
        self.custom_lines.clear()
        self.texts.clear()
        self._undo.clear()
        self._redo.clear()
        return errors

    # ---- 그룹 속성 접근 -----------------------------------------------------
    def group_rel(self, group):
        return self.g[group].get("rel", self.reliability)

    def group_lot(self, group):
        return self.g[group].get("lot", group)

    def readouts(self, group):
        return self.g[group]["readouts"]

    def columns(self, group):
        return self.g[group]["columns"]

    def samples(self, group):
        return self.g[group]["samples"]

    # ---- 값 조회 -----------------------------------------------------------
    def value(self, group, readout, col, sample):
        if (group, readout, col, sample) in self.deleted:
            return None
        return self.g[group]["data"].get(readout, {}).get(sample, {}).get(col)

    def series(self, group, readout, col):
        xs, ys = [], []
        for s in self.samples(group):
            xs.append(s)
            v = self.value(group, readout, col, s)
            ys.append(math.nan if v is None else v)
        return xs, ys

    def delta_series(self, group, readout, col):
        """초기 Read-out 대비 변화율(%). 기준은 항상 그룹의 첫 Read-out.
        초기값이 0이거나 결측이면 해당 점은 NaN."""
        base = self.readouts(group)[0]
        xs, ys = [], []
        for s in self.samples(group):
            xs.append(s)
            v0 = self.value(group, base, col, s)
            v = self.value(group, readout, col, s)
            if v0 is None or v is None or v0 == 0:
                ys.append(math.nan)
            else:
                ys.append((v - v0) / v0 * 100.0)
        return xs, ys

    def box_values(self, group, readout, col):
        return [v for s in self.samples(group)
                if (v := self.value(group, readout, col, s)) is not None]

    # ---- diff 기반 Undo/Redo ------------------------------------------------
    def _apply(self, action, forward=True):
        kind = action[0]
        if kind == "del":
            _, keys = action
            if forward:
                self.deleted.update(keys)
            else:
                self.deleted.difference_update(keys)
        elif kind == "color":
            _, key, old, new = action
            c = new if forward else old
            if c is None:
                self.color_over.pop(key, None)
            else:
                self.color_over[key] = c
        elif kind == "ylim":
            _, key, old, new = action
            v = new if forward else old
            if v is None:
                self.ylim.pop(key, None)
            else:
                self.ylim[key] = v
        elif kind == "limits":
            _, key, old, new = action
            v = new if forward else old
            if not v:
                self.limits.pop(key, None)
            else:
                self.limits[key] = v
        elif kind == "cline":
            _, key, old, new = action
            v = new if forward else old
            if not v:
                self.custom_lines.pop(key, None)
            else:
                self.custom_lines[key] = v
        elif kind == "text":
            _, key, old, new = action
            v = new if forward else old
            if not v:
                self.texts.pop(key, None)
            else:
                self.texts[key] = v

    def do(self, action):
        self._apply(action, True)
        self._undo.append(action)
        self._redo.clear()

    def undo(self):
        if not self._undo:
            return False
        a = self._undo.pop()
        self._apply(a, False)
        self._redo.append(a)
        return True

    def redo(self):
        if not self._redo:
            return False
        a = self._redo.pop()
        self._apply(a, True)
        self._undo.append(a)
        return True

    # ---- 편집 액션 -----------------------------------------------------------
    def delete_point(self, group, readout, col, sample):
        self.do(("del", frozenset({(group, readout, col, sample)})))

    def delete_sample_all_readouts(self, group, col, sample):
        keys = frozenset((group, r, col, sample) for r in self.readouts(group))
        self.do(("del", keys))

    def set_color(self, group, readout, col, sample, color):
        key = (group, readout, col, sample)
        self.do(("color", key, self.color_over.get(key), color))

    def set_limits(self, group, col, high, low):
        key = (group, col)
        old = dict(self.limits.get(key, {}))
        new = {}
        if high is not None:
            new["high"] = high
        if low is not None:
            new["low"] = low
        self.do(("limits", key, old, new))

    def set_custom_lines(self, group, col, kind, lines):
        key = (group, col, kind)
        old = list(self.custom_lines.get(key, []))
        self.do(("cline", key, old, list(lines)))

    def set_texts(self, group, col, kind, texts):
        key = (group, col, kind)
        old = list(self.texts.get(key, []))
        self.do(("text", key, old, list(texts)))

    def set_ylim(self, group, col, ymin, ymax):
        key = (group, col)
        self.do(("ylim", key, self.ylim.get(key),
                 None if ymin is None else (ymin, ymax)))

    # ---- 통계 -----------------------------------------------------------------
    def stats(self, group, readout, col):
        vals = self.box_values(group, readout, col)
        n = len(vals)
        if n == 0:
            return dict(SS=0, Min=math.nan, Max=math.nan, AVG=math.nan, STD=math.nan)
        avg = sum(vals) / n
        std = (sum((v - avg) ** 2 for v in vals) / (n - 1)) ** 0.5 if n > 1 else 0.0
        return dict(SS=n, Min=min(vals), Max=max(vals), AVG=avg, STD=std)


# ============================================================================
# 4. 그래프 렌더링 (화면/PDF 공용 파이프라인) — 모두 (group, col) 단위
# ============================================================================
LIMIT_COLOR = "#d62728"


def _collect_pts(ax):
    pts_x, pts_y = [], []
    for ln in ax.get_lines():
        if ln.get_marker() in ("o", "*"):
            for x, y in zip(ln.get_xdata(), ln.get_ydata()):
                try:
                    if not math.isnan(float(y)):
                        pts_x.append(float(x))
                        pts_y.append(float(y))
                except (TypeError, ValueError):
                    pass
    return pts_x, pts_y


def _place_hline_label(ax, v, name, color, pts_x, pts_y, occupied):
    """가로선 이름: 선 바로 위를 좌우로 훑어 data/기존 이름과 겹치지 않는 곳에.
    실패 시 빈 공간 이동 + 연결선."""
    y0, y1 = ax.get_ylim()
    x0, x1 = ax.get_xlim()
    yr = (y1 - y0) or 1.0
    xr = (x1 - x0) or 1.0
    dy = 0.05 * yr
    ty = v + 0.010 * yr
    est_w = 0.028 * xr * max(len(name), 1)  # 대략적 텍스트 폭
    if 0.02 < (v - y0) / yr < 0.965:
        for xf in [0.06, 0.16, 0.26, 0.36, 0.46, 0.56, 0.66, 0.76, 0.86]:
            cx = x0 + xf * xr
            box = (cx, cx + est_w, ty, ty + 0.06 * yr)
            if any(not (box[1] < o[0] or box[0] > o[1] or
                        box[3] < o[2] or box[2] > o[3]) for o in occupied):
                continue
            crowded = any(abs(px - cx) < 0.07 * xr and
                          v - 0.5 * dy <= py <= v + 2.2 * dy
                          for px, py in zip(pts_x, pts_y))
            if not crowded:
                ax.text(cx, ty, name, ha="left", va="bottom",
                        fontsize=7, color=color, zorder=6)
                occupied.append(box)
                return
    # 폴백: 좌측 빈 공간 + 연결선
    above = sum(1 for y in pts_y if v < y <= v + 2 * dy)
    below = sum(1 for y in pts_y if v - 2 * dy <= y < v)
    ty2 = v + 1.8 * dy if above <= below else v - 1.8 * dy
    ty2 = min(max(ty2, y0 + 0.04 * yr), y1 - 0.04 * yr)
    tx = x0 + 0.02 * xr
    ax.annotate(name, xy=(x0 + 0.10 * xr, v), xytext=(tx, ty2),
                ha="left", va="center", fontsize=7, color=color, zorder=6,
                arrowprops=dict(arrowstyle="-", color=color, lw=0.7, alpha=0.9))
    occupied.append((tx, tx + est_w, ty2 - 0.03 * yr, ty2 + 0.03 * yr))


def _place_vline_label(ax, v, name, color, pts_x, pts_y, occupied):
    """세로선 이름: 선을 따라 상하로 훑어 겹치지 않는 곳에. 실패 시 이동+연결선."""
    y0, y1 = ax.get_ylim()
    x0, x1 = ax.get_xlim()
    yr = (y1 - y0) or 1.0
    xr = (x1 - x0) or 1.0
    est_w = 0.028 * xr * max(len(name), 1)
    for yf in [0.90, 0.78, 0.66, 0.54, 0.42, 0.30]:
        cy = y0 + yf * yr
        box = (v + 0.008 * xr, v + 0.008 * xr + est_w,
               cy - 0.035 * yr, cy + 0.035 * yr)
        if any(not (box[1] < o[0] or box[0] > o[1] or
                    box[3] < o[2] or box[2] > o[3]) for o in occupied):
            continue
        crowded = any(abs(px - v) < 0.05 * xr and abs(py - cy) < 0.07 * yr
                      for px, py in zip(pts_x, pts_y))
        if not crowded:
            ax.text(v + 0.008 * xr, cy, name, ha="left", va="center",
                    fontsize=7, color=color, zorder=6)
            occupied.append(box)
            return
    dx = 0.05 * xr
    tx = min(max(v + 2 * dx, x0 + 0.04 * xr), x1 - 0.04 * xr)
    ty = y1 - 0.12 * yr
    ax.annotate(name, xy=(v, ty - 0.06 * yr), xytext=(tx, ty),
                ha="center", va="bottom", fontsize=7, color=color, zorder=6,
                arrowprops=dict(arrowstyle="-", color=color, lw=0.7, alpha=0.9))
    occupied.append((tx - est_w / 2, tx + est_w / 2, ty, ty + 0.06 * yr))


def draw_overlays(ax, model, group, col, kind):
    """High/Low limit(빨간 점선) + 임의의 선 + 임의 텍스트 렌더링.
    이름들은 서로/데이터와 겹치지 않게 배치."""
    pts_x, pts_y = _collect_pts(ax)
    occupied = []
    # 1) High/Low limit — Read-out('line')과 Box('box')에만 표시 (Delta 제외)
    lim = model.limits.get((group, col), {})
    if kind in ("line", "box") and lim:
        for key, name in (("high", "High limit"), ("low", "Low limit")):
            if lim.get(key) is None:
                continue
            v = lim[key]
            ax.axhline(v, color=LIMIT_COLOR, ls=":", lw=1.2, zorder=4)
            _place_hline_label(ax, v, name, LIMIT_COLOR, pts_x, pts_y, occupied)
    # 2) 임의의 선
    for cl in model.custom_lines.get((group, col, kind), []):
        c = cl.get("color") or "#1f77b4"
        st = cl.get("style", ":")
        nm = cl.get("name", "")
        v = cl["value"]
        if cl["axis"] == "y":
            ax.axhline(v, color=c, ls=st, lw=1.2, zorder=4)
            if nm:
                _place_hline_label(ax, v, nm, c, pts_x, pts_y, occupied)
        else:
            ax.axvline(v, color=c, ls=st, lw=1.2, zorder=4)
            if nm:
                _place_vline_label(ax, v, nm, c, pts_x, pts_y, occupied)
    # 3) 임의 텍스트 (드래그 이동 지원 — GUI에서 picker 연결)
    artists = []
    for i, tx in enumerate(model.texts.get((group, col, kind), [])):
        t = ax.text(tx["x"], tx["y"], tx["text"],
                    fontsize=tx.get("size", 10), color=tx.get("color", "black"),
                    fontfamily=["Arial", "DejaVu Sans"], zorder=7, picker=True)
        t._user_text_index = i
        t._user_text_kind = kind
        artists.append(t)
    return artists


def limit_out(model, group, col, v):
    """값 v가 High/Low limit 밖인지."""
    lim = model.limits.get((group, col), {})
    if not lim or v is None:
        return False
    hi = lim.get("high")
    lo = lim.get("low")
    return (hi is not None and v > hi) or (lo is not None and v < lo)


def readout_color(model, group, readout):
    return READOUT_COLORS[model.readouts(group).index(readout) % len(READOUT_COLORS)]


def _apply_x_axis(ax, model, group):
    samples = model.samples(group)
    if samples:
        ax.set_xlim(min(samples) - 1, max(samples) + 1)
        even = [s for s in samples if s % 2 == 0]
        ax.set_xticks(even)
        ax.xaxis.set_minor_locator(MultipleLocator(1))
    ax.tick_params(labelsize=7)


def graph_title(model, group, col):
    """제목 형식: 신뢰성: LOT번호, ITEM [단위] @조건 (Retest면 RETEST_ 접두)"""
    return f"{model.group_rel(group)}: {model.group_lot(group)}, {col_title(col)}"


def draw_line(ax, model, group, col, picker=False, screen=False):
    """Read-out graph: X=Sample(짝수 Label/홀수 Minor Tick), Read-out별 색상.
    limit 밖 data는 원래 색 유지 + 굵은 * 마커."""
    artists = {}
    base_ms = 4
    for r in model.readouts(group):
        c = readout_color(model, group, r)
        xs, ys = model.series(group, r, col)
        kw = dict(lw=1, color=c, label=r)
        if picker:
            kw["picker"] = 5
        line, = ax.plot(xs, ys, **kw)
        artists[r] = line
        in_x, in_y, out_x, out_y = [], [], [], []
        for x, v in zip(xs, ys):
            if math.isnan(v):
                continue
            if limit_out(model, group, col, v):
                out_x.append(x)
                out_y.append(v)
            else:
                in_x.append(x)
                in_y.append(v)
        ax.plot(in_x, in_y, marker="o", ms=base_ms, ls="none", color=c)
        if out_x:
            ax.plot(out_x, out_y, marker="*", ms=base_ms + 3, mew=1.1,
                    ls="none", color=c, zorder=4)
        for s, v in zip(xs, ys):
            oc = model.color_over.get((group, r, col, s))
            if oc and not math.isnan(v):
                ax.plot([s], [v], marker="^", ms=6, color=oc,
                        markeredgecolor="black", markeredgewidth=0.6,
                        ls="none", zorder=5)
    ax.set_title(graph_title(model, group, col), fontsize=9, loc="left")
    ax.set_xlabel("Sample No.", fontsize=8)
    mu = re.search(r"\(([^()]*)\)\s*(?:#\d+)?$", col)
    unit = f" ({mu.group(1)})" if mu else ""
    ax.set_ylabel(col.split("@")[0] + unit, fontsize=8)
    _apply_x_axis(ax, model, group)
    if (group, col) in model.ylim:
        ax.set_ylim(*model.ylim[(group, col)])
    else:
        # 기본: 모든 Read-out 그래프는 Y축 0 ~ (데이터 max의 110%)
        vmax = None
        for _ln in ax.get_lines():
            for _y in _ln.get_ydata():
                try:
                    fy = float(_y)
                except (TypeError, ValueError):
                    continue
                if not math.isnan(fy) and (vmax is None or fy > vmax):
                    vmax = fy
        if vmax is not None and vmax > 0:
            ax.set_ylim(0, vmax * 1.10)
        else:
            ax.set_ylim(bottom=0)
    # legend: 축 바깥 위쪽 가로 배치 → 그래프와 겹치지 않음
    ax.legend(fontsize=9 if screen else 6,
              ncol=max(1, len(model.readouts(group))),
              loc="lower right", bbox_to_anchor=(1.0, 1.0),
              borderaxespad=0.0, frameon=True, handletextpad=0.4,
              columnspacing=0.8)
    text_artists = draw_overlays(ax, model, group, col, "line")
    ax.grid(True, alpha=0.3)
    artists["__texts__"] = text_artists
    return artists


def delta_title(model, group, col):
    """Delta 제목: 단위 위치에 일괄 [%] 적용."""
    t = f"{model.group_rel(group)}: {model.group_lot(group)}, {col_title(col)}"
    return re.sub(r"\[[^\[\]]*\]", "[%]", t, count=1)


def draw_delta(ax, model, group, col, screen=False, picker=False):
    """Delta % graph: 초기 Read-out 기준 변화율. layout은 Read-out graph와 동일."""
    artists = {}
    for r in model.readouts(group)[1:]:
        c = readout_color(model, group, r)
        xs, ys = model.delta_series(group, r, col)
        kw = dict(marker="o", ms=4, lw=1, color=c, label=r)
        if picker:
            kw["picker"] = 5
        line, = ax.plot(xs, ys, **kw)
        artists[r] = line
        for s, v in zip(xs, ys):
            oc = model.color_over.get((group, r, col, s))
            if oc and not math.isnan(v):
                ax.plot([s], [v], marker="^", ms=6, color=oc,
                        markeredgecolor="black", markeredgewidth=0.6,
                        ls="none", zorder=5)
    ax.set_title(delta_title(model, group, col), fontsize=9, loc="left")
    ax.set_xlabel("Sample No.", fontsize=8)
    ax.set_ylabel("Delta (%)", fontsize=8)
    _apply_x_axis(ax, model, group)
    ax.axhline(0, color="gray", lw=0.8, ls="--", alpha=0.7)
    if len(model.readouts(group)) > 1:
        ax.legend(fontsize=9 if screen else 6,
                  ncol=max(1, len(model.readouts(group)) - 1),
                  loc="lower right", bbox_to_anchor=(1.0, 1.0),
                  borderaxespad=0.0, frameon=True, handletextpad=0.4,
                  columnspacing=0.8)
    else:
        ax.text(0.5, 0.5, "Read-out 1개 — Delta 없음", transform=ax.transAxes,
                ha="center", va="center", fontsize=9, color="gray")
    text_artists = draw_overlays(ax, model, group, col, "delta")
    ax.grid(True, alpha=0.3)
    artists["__texts__"] = text_artists
    return artists


def draw_box(ax, model, group, col, picker=False, stats_table=True):
    """Box plot: X=Read-out, Read-out graph와 동일 색상, Median/Outlier 표시.
    stats_table=True면 각 Read-out 박스 x위치에 맞춰 통계 텍스트 표시."""
    readouts = model.readouts(group)
    groups_vals = [model.box_values(group, r, col) for r in readouts]
    # Matplotlib 3.9+에서 'labels' → 'tick_labels'로 변경됨 (구버전 폴백 포함)
    try:
        bp = ax.boxplot(groups_vals, tick_labels=readouts, patch_artist=True,
                        showfliers=True, medianprops=dict(color="black"))
    except TypeError:
        bp = ax.boxplot(groups_vals, labels=readouts, patch_artist=True,
                        showfliers=True, medianprops=dict(color="black"))
    for patch, r in zip(bp["boxes"], readouts):
        c = readout_color(model, group, r)
        patch.set_facecolor(c)
        patch.set_alpha(0.5)
    star_artists = []
    for i, (fl, r) in enumerate(zip(bp["fliers"], readouts)):
        c = readout_color(model, group, r)
        yd = list(fl.get_ydata())
        in_l = [y for y in yd if not limit_out(model, group, col, y)]
        out_l = [y for y in yd if limit_out(model, group, col, y)]
        fl.set_xdata([i + 1] * len(in_l))
        fl.set_ydata(in_l)
        fl.set(marker="o", markerfacecolor=c,
               markeredgecolor="black", markersize=3)
        if picker:
            fl.set_picker(5)
        if out_l:  # limit 밖 outlier는 굵은 * (원래 색 유지)
            st, = ax.plot([i + 1] * len(out_l), out_l, marker="*", ms=6,
                          mew=1.1, ls="none", color=c, zorder=4)
            if picker:
                st.set_picker(5)
            st._star_readout = r
            star_artists.append(st)
    # 색 변경된 점은 삼각형으로 해당 Read-out 위치에 표시 (Line과 동일 규칙)
    for (gr, r, c, s), oc in model.color_over.items():
        if gr != group or c != col:
            continue
        v = model.value(group, r, c, s)
        if v is not None and r in readouts:
            ax.plot([readouts.index(r) + 1], [v], marker="^", ms=6,
                    color=oc, markeredgecolor="black", markeredgewidth=0.6,
                    ls="none", zorder=5)
    ax.set_title(graph_title(model, group, col), fontsize=9, loc="left")
    ax.tick_params(labelsize=7)
    if (group, col) in model.ylim:
        ax.set_ylim(*model.ylim[(group, col)])
    text_artists = draw_overlays(ax, model, group, col, "box")
    ax.grid(True, alpha=0.3)

    if stats_table:
        # x축 tick 라벨은 숨기고, Read-out 이름을 통계 머리행으로 대신 표시하여
        # 어떤 페이지 layout에서도 라벨/통계 겹침이 발생하지 않도록 함
        trans = mtransforms.blended_transform_factory(ax.transData, ax.transAxes)
        ax.tick_params(axis="x", labelbottom=False)
        y0, dy, fs = -0.145, 0.075, 7
        for i, r in enumerate(readouts):
            ax.text(i + 1, y0, r, transform=trans,
                    ha="center", va="center", fontsize=fs, fontweight="bold")
        rows = [("S/S", "SS"), ("Min", "Min"), ("Max", "Max"),
                ("AVG", "AVG"), ("STD", "STD")]
        for k, (label, key) in enumerate(rows):
            y = y0 - dy * (k + 1)
            ax.text(-0.01, y, label, transform=ax.transAxes,
                    ha="right", va="center", fontsize=fs, fontweight="bold")
            for i, r in enumerate(readouts):
                st = model.stats(group, r, col)
                v = st[key]
                txt = str(v) if key == "SS" else f"{v:.4g}"
                ax.text(i + 1, y, txt, transform=trans,
                        ha="center", va="center", fontsize=fs)
    else:
        ax.set_xlabel("Read-out", fontsize=8)
    bp["stars"] = star_artists
    bp["texts"] = text_artists
    return bp


# ============================================================================
# 5. PDF 출력 — 그룹별 섹션: Read-out graph → Delta % graph → Box plot
#    페이지: PPT 16:9 가로. Line/Delta 1열x3줄, Box 3개x2줄.
# ============================================================================
PPT_LANDSCAPE = (13.33, 7.5)  # 16:9 PPT 슬라이드 가로 방향


def export_pdf(model, pairs, path, progress_cb=None):
    """pairs: [(lot, col)] 선택 목록. Lot 순서대로 섹션 생성.
    Lot별 구성: item마다 [Read-out graph(위) + Delta % graph(아래)] 1페이지 쌍
                → 이어서 Box plot(3개 x 2줄) item 순서대로."""
    per_group = {}
    for g, c in pairs:
        per_group.setdefault(g, []).append(c)
    order = [g for g in model.groups if g in per_group]

    total = sum(math.ceil(len(per_group[g]) / 2)
                + math.ceil(len(per_group[g]) / 6) for g in order)
    done = 0
    with PdfPages(path) as pdf:
        for g in order:
            cols = per_group[g]
            head = f"{model.group_rel(g)}: {model.group_lot(g)}"
            # 1) item별 Read-out + Delta % 쌍 — 1페이지 1줄 1개 x 4줄 (item 2개분)
            for i in range(0, len(cols), 2):
                fig = Figure(figsize=PPT_LANDSCAPE)
                fig.suptitle(f"{head} — Read-out & Delta %", fontsize=12)
                for k in range(2):  # 마지막 페이지도 4칸 유지 → 동일 크기
                    ax1 = fig.add_subplot(4, 1, 2 * k + 1)
                    ax2 = fig.add_subplot(4, 1, 2 * k + 2)
                    if i + k < len(cols):
                        c = cols[i + k]
                        draw_line(ax1, model, g, c)
                        draw_delta(ax2, model, g, c)
                    else:
                        ax1.axis("off")
                        ax2.axis("off")
                fig.tight_layout(rect=(0, 0, 1, 0.96))
                pdf.savefig(fig)
                done += 1
                if progress_cb:
                    progress_cb(done, total)
            # 2) Box plot: 3개 x 2줄, item 순서대로
            for i in range(0, len(cols), 6):
                fig = Figure(figsize=PPT_LANDSCAPE)
                fig.suptitle(f"{head} — Box plot", fontsize=12)
                for k in range(6):
                    ax = fig.add_subplot(2, 3, k + 1)
                    if i + k < len(cols):
                        draw_box(ax, model, g, cols[i + k], stats_table=True)
                    else:
                        ax.axis("off")
                fig.subplots_adjust(top=0.90, bottom=0.15, left=0.07, right=0.98,
                                    hspace=0.85, wspace=0.35)
                pdf.savefig(fig)
                done += 1
                if progress_cb:
                    progress_cb(done, total)


# ============================================================================
# 6. GUI
# ============================================================================
BaseTk = TkinterDnD.Tk if HAS_DND else tk.Tk


class App(BaseTk):
    def __init__(self):
        super().__init__()
        self.title("Discrete Reliability Data Analyzer")
        self.geometry("1200x800")
        center_window(self, 1200, 800)
        self.model = DataModel()
        self.files = []
        self.selected = []   # [(group, col)]
        self.cur_idx = 0
        self.box_page = 0
        self._build_start()

    def _reset_all(self):
        """진행 내용을 모두 초기화하고 처음 화면으로."""
        if not messagebox.askyesno("처음으로", "진행하던 내용을 모두 지우고\n처음 화면으로 돌아갈까요?",
                                   parent=self):
            return
        self.model = DataModel()
        self.files = []
        self.selected = []
        self.cur_idx = 0
        self._build_start()

    # ---- 화면 1: 파일 선택 -------------------------------------------------
    def _build_start(self):
        for w in self.winfo_children():
            w.destroy()
        frm = ttk.Frame(self, padding=20)
        frm.pack(fill="both", expand=True)
        ttk.Label(frm, text="Discrete Reliability Data Analyzer",
                  font=("", 16, "bold")).pack(pady=10)
        ttk.Label(frm, text="파일 이름은 신뢰성명 + Lot번호 + Read-out 형식이어야 합니다.\n"
                            "예: HTRB_Lot1_0hr.csv, HTBG+_Lot2_500cyc.xlsx (구분자: _ 또는 공백)\n"
                            "서로 다른 신뢰성/Lot은 자동으로 구분되어 각각 분석됩니다.").pack(pady=5)

        drop = tk.Label(frm, text="여기에 파일을 Drag && Drop 하세요"
                        if HAS_DND else "Browse 버튼으로 파일을 선택하세요",
                        relief="ridge", height=6, bg="#f0f0f0")
        drop.pack(fill="x", pady=10)
        if HAS_DND:
            drop.drop_target_register(DND_FILES)
            drop.dnd_bind("<<Drop>>", lambda e: self._add_files(self.tk.splitlist(e.data)))

        ttk.Button(frm, text="Browse...", command=self._browse).pack()
        self.file_list = tk.Listbox(frm, height=8)
        self.file_list.pack(fill="both", expand=True, pady=10)
        btns = ttk.Frame(frm)
        btns.pack()
        ttk.Button(btns, text="선택 제거", command=self._remove_file).pack(side="left", padx=5)
        ttk.Button(btns, text="다음 (파일 읽기)", command=self._load_files).pack(side="left", padx=5)

    def _browse(self):
        paths = filedialog.askopenfilenames(
            filetypes=[("Data files", "*.csv *.xlsx *.xlsm"), ("All", "*.*")])
        self._add_files(paths)

    def _add_files(self, paths):
        for p in paths:
            p = p.strip("{}")
            if p and p not in self.files:
                self.files.append(p)
                self.file_list.insert("end", os.path.basename(p))

    def _remove_file(self):
        for i in reversed(self.file_list.curselection()):
            self.file_list.delete(i)
            del self.files[i]

    def _load_files(self):
        if not self.files:
            messagebox.showwarning("알림", "파일을 먼저 선택하세요.", parent=self)
            return
        errors = self.model.load(self.files)
        if errors:
            messagebox.showerror("파일 오류", "\n\n".join(errors), parent=self)
        if not self.model.groups:
            return
        self._build_param_select()

    # ---- 화면 2: Parameter 선택 (그룹 | Parameter) ----------------------------
    def _build_param_select(self):
        for w in self.winfo_children():
            w.destroy()
        frm = ttk.Frame(self, padding=20)
        frm.pack(fill="both", expand=True)
        head = [f"신뢰성: {self.model.reliability}"]
        for g in self.model.groups:
            head.append(f"{g} (Read-out: {', '.join(self.model.readouts(g))}, "
                        f"Sample: {len(self.model.samples(g))})")
        ttk.Label(frm, text="\n".join(head), font=("", 10, "bold"),
                  justify="left").pack(pady=5)
        ttk.Label(frm, text="분석할 Parameter를 선택하세요 (Ctrl/Shift 다중 선택)").pack()
        self.param_lb = tk.Listbox(frm, selectmode="extended")
        self._pairs = []
        multi = len(self.model.groups) > 1
        for g in self.model.groups:
            for c in self.model.columns(g):
                self._pairs.append((g, c))
                label = col_title(c)
                self.param_lb.insert("end", f"{g} | {label}" if multi else label)
        self.param_lb.pack(fill="both", expand=True, pady=10)
        btns = ttk.Frame(frm)
        btns.pack()
        ttk.Button(btns, text="전체 선택",
                   command=lambda: self.param_lb.select_set(0, "end")).pack(side="left", padx=5)
        ttk.Button(btns, text="분석 시작", command=self._analyze).pack(side="left", padx=5)
        ttk.Button(btns, text="← 파일 다시 선택", command=self._build_start).pack(side="left", padx=5)
        ttk.Button(btns, text="⟲ 처음으로 (Reset)", command=self._reset_all).pack(side="left", padx=5)
        self.pbar = ttk.Progressbar(frm, mode="determinate")
        self.pbar.pack(fill="x", pady=5)

    def _analyze(self):
        sel = self.param_lb.curselection()
        if not sel:
            messagebox.showwarning("알림", "Parameter를 선택하세요.", parent=self)
            return
        self.selected = [self._pairs[i] for i in sel]
        self.cur_idx = 0
        self.pbar["maximum"] = len(self.selected)
        self.pbar["value"] = len(self.selected)
        self.update_idletasks()
        self._build_graphs()

    # ---- 화면 3: 그래프 -------------------------------------------------------
    def _label(self, pair):
        g, c = pair
        label = col_title(c)
        return f"{g} | {label}" if len(self.model.groups) > 1 else label

    def _build_graphs(self):
        for w in self.winfo_children():
            w.destroy()
        top = ttk.Frame(self, padding=5)
        top.pack(fill="x")
        ttk.Button(top, text="← Parameter", command=self._build_param_select).pack(side="left")
        ttk.Button(top, text="◀ 이전", command=lambda: self._nav(-1)).pack(side="left", padx=3)
        self.param_var = tk.StringVar()
        self._labels = [self._label(p) for p in self.selected]
        cb = ttk.Combobox(top, textvariable=self.param_var,
                          values=self._labels, width=55, state="readonly")
        cb.pack(side="left", padx=3)
        cb.bind("<<ComboboxSelected>>",
                lambda e: self._goto(self._labels.index(self.param_var.get())))
        ttk.Button(top, text="다음 ▶", command=lambda: self._nav(1)).pack(side="left", padx=3)
        ttk.Separator(top, orient="vertical").pack(side="left", fill="y", padx=8)
        ttk.Button(top, text="Y축 Min/Max", command=self._set_ylim).pack(side="left", padx=3)
        ttk.Button(top, text="Limit", command=self._limit_dialog).pack(side="left", padx=3)
        ttk.Button(top, text="임의선", command=self._custom_line_dialog).pack(side="left", padx=3)
        ttk.Button(top, text="텍스트", command=self._text_dialog).pack(side="left", padx=3)
        ttk.Button(top, text="Undo", command=self._undo).pack(side="left", padx=3)
        ttk.Button(top, text="Redo", command=self._redo).pack(side="left", padx=3)
        ttk.Button(top, text="Export PDF", command=self._export_pdf).pack(side="right", padx=3)
        ttk.Button(top, text="⟲ 처음으로", command=self._reset_all).pack(side="right", padx=3)

        self.nb = ttk.Notebook(self)
        self.nb.pack(fill="both", expand=True)
        self.line_tab = ttk.Frame(self.nb)   # Read-out(위) + Delta %(아래) 통합 탭
        self.box_tab = ttk.Frame(self.nb)
        self.nb.add(self.line_tab, text="Read-out & Delta % graph")
        self.nb.add(self.box_tab, text="Box plot")

        self.line_fig = Figure(figsize=(10, 7))
        self.line_canvas = FigureCanvasTkAgg(self.line_fig, self.line_tab)
        self.line_canvas.get_tk_widget().pack(fill="both", expand=True)
        NavigationToolbar2Tk(self.line_canvas, self.line_tab)
        self.line_canvas.mpl_connect("pick_event", self._on_pick_line)
        self.line_canvas.mpl_connect("motion_notify_event", self._on_text_motion)
        self.line_canvas.mpl_connect("button_release_event", self._on_text_release)

        ctrl = ttk.Frame(self.box_tab)
        ctrl.pack(fill="x")
        ttk.Button(ctrl, text="◀", width=3,
                   command=lambda: self._box_nav(-1)).pack(side="left", padx=(8, 2))
        self.box_page_lbl = ttk.Label(ctrl, text="")
        self.box_page_lbl.pack(side="left")
        ttk.Button(ctrl, text="▶", width=3,
                   command=lambda: self._box_nav(1)).pack(side="left", padx=2)

        self.box_fig = Figure(figsize=(10, 5))
        self.box_canvas = FigureCanvasTkAgg(self.box_fig, self.box_tab)
        self.box_canvas.get_tk_widget().pack(fill="both", expand=True)
        NavigationToolbar2Tk(self.box_canvas, self.box_tab)
        self.box_canvas.mpl_connect("pick_event", self._on_pick_box)
        self.box_canvas.mpl_connect("motion_notify_event", self._on_text_motion)
        self.box_canvas.mpl_connect("button_release_event", self._on_text_release)

        self._goto(0)

    def _nav(self, d):
        nxt = self.cur_idx + d
        if nxt >= len(self.selected):
            messagebox.showinfo("알림", "마지막 item입니다.", parent=self)
            return
        if nxt < 0:
            messagebox.showinfo("알림", "처음 item입니다.", parent=self)
            return
        self._goto(nxt)

    def _goto(self, idx):
        self.cur_idx = idx
        self._sel_point = None
        group, col = self.selected[idx]
        same_group = [p for p in self.selected if p[0] == group]
        self.box_page = same_group.index((group, col)) // 4
        self.param_var.set(self._labels[idx])
        self._redraw()

    def _redraw(self):
        group, col = self.selected[self.cur_idx]

        self.line_fig.clear()
        self._drag_text = None
        ax = self.line_fig.add_subplot(2, 1, 1)
        self._line_artists = draw_line(ax, self.model, group, col,
                                       picker=True, screen=True)
        axd = self.line_fig.add_subplot(2, 1, 2)
        self._delta_artists = draw_delta(axd, self.model, group, col,
                                         screen=True, picker=True)
        self._ax_line, self._ax_delta = ax, axd
        # 선택 강조: 양쪽 그래프에서 동일 data 마커 표시
        sel = getattr(self, "_sel_point", None)
        if sel is not None:
            s_r, s_s = sel
            # Read-out 그래프
            v = self.model.value(group, s_r, col, s_s)
            if v is not None:
                ax.plot([s_s], [v], marker="o", ms=10, mfc="none",
                        mec="#ff9900", mew=1.8, ls="none", zorder=8)
            # Delta 그래프 (기준 read-out은 delta에 없음)
            if s_r != self.model.readouts(group)[0]:
                xs, ys = self.model.delta_series(group, s_r, col)
                for x, dv in zip(xs, ys):
                    if x == s_s and not math.isnan(dv):
                        axd.plot([x], [dv], marker="o", ms=10, mfc="none",
                                 mec="#ff9900", mew=1.8, ls="none", zorder=8)
        self.line_fig.tight_layout()
        self.line_canvas.draw()

        self._redraw_box()

    def _box_nav(self, d):
        group, _ = self.selected[self.cur_idx]
        same_group = [p for p in self.selected if p[0] == group]
        pages = max(1, math.ceil(len(same_group) / 4))
        nxt = self.box_page + d
        if nxt < 0 or nxt >= pages:
            return
        self.box_page = nxt
        self._redraw_box()

    def _redraw_box(self):
        group, col = self.selected[self.cur_idx]
        self.box_fig.clear()
        self._box_fliers = {}  # flier artist → (group, readout, colname)
        same_group = [p for p in self.selected if p[0] == group]
        pages = max(1, math.ceil(len(same_group) / 4))
        self.box_page = min(self.box_page, pages - 1)
        self.box_page_lbl.config(text=f"{self.box_page + 1}/{pages}")
        page_items = same_group[self.box_page * 4:(self.box_page + 1) * 4]
        for k, (g2, c2) in enumerate(page_items):
            ax2 = self.box_fig.add_subplot(2, 2, k + 1)
            bp = draw_box(ax2, self.model, g2, c2, picker=True, stats_table=True)
            for fl, r in zip(bp["fliers"], self.model.readouts(g2)):
                self._box_fliers[fl] = (g2, r, c2)
            for st in bp.get("stars", []):
                self._box_fliers[st] = (g2, st._star_readout, c2)
            if c2 == col:
                ax2.set_title(graph_title(self.model, g2, c2),
                              fontsize=9, fontweight="bold")
        self.box_fig.subplots_adjust(top=0.93, bottom=0.16, left=0.10,
                                     right=0.97, hspace=0.85, wspace=0.35)
        self.box_canvas.draw()

    # ---- 편집 --------------------------------------------------------------
    def _on_pick_line(self, event):
        # 텍스트 드래그 시작
        if hasattr(event.artist, "_user_text_index"):
            self._drag_text = event.artist
            self._drag_kind = event.artist._user_text_kind
            return
        group, col = self.selected[self.cur_idx]
        artist = event.artist
        readout = None
        for d in (getattr(self, "_line_artists", {}) or {},
                  getattr(self, "_delta_artists", {}) or {}):
            for r, ln in d.items():
                if r != "__texts__" and ln is artist:
                    readout = r
                    break
            if readout:
                break
        if readout is None or not len(event.ind):
            return
        sample = self.model.samples(group)[event.ind[0]]
        self._sel_point = (readout, sample)
        self._redraw()  # 양쪽 그래프 동일 마커 강조
        self._point_menu(group, readout, col, sample)

    def _on_text_motion(self, event):
        t = getattr(self, "_drag_text", None)
        if t is None or event.xdata is None or event.ydata is None:
            return
        t.set_position((event.xdata, event.ydata))
        t.figure.canvas.draw_idle()

    def _on_text_release(self, event):
        t = getattr(self, "_drag_text", None)
        if t is None:
            return
        self._drag_text = None
        if event.xdata is None or event.ydata is None:
            return
        group, col = self.selected[self.cur_idx]
        kind = self._drag_kind
        texts = list(self.model.texts.get((group, col, kind), []))
        i = t._user_text_index
        if 0 <= i < len(texts):
            texts[i] = dict(texts[i], x=float(event.xdata), y=float(event.ydata))
            self.model.set_texts(group, col, kind, texts)
        self._redraw()

    def _on_pick_box(self, event):
        if hasattr(event.artist, "_user_text_index"):
            self._drag_text = event.artist
            self._drag_kind = event.artist._user_text_kind
            return
        info = getattr(self, "_box_fliers", {}).get(event.artist)
        if info is None or not len(event.ind):
            return
        group, readout, col = info
        yval = event.artist.get_ydata()[event.ind[0]]
        sample = None
        best = float("inf")
        for s in self.model.samples(group):
            v = self.model.value(group, readout, col, s)
            if v is None:
                continue
            d = abs(v - yval)
            if d < best:
                best, sample = d, s
        if sample is None:
            return
        self._point_menu(group, readout, col, sample)

    def _point_menu(self, group, readout, col, sample):
        m = tk.Menu(self, tearoff=0)
        m.add_command(label=f"{group} | Sample {sample} @ {readout}", state="disabled")
        m.add_separator()
        m.add_command(label="Marker 색 변경 (삼각형 표시)",
                      command=lambda: self._change_color(group, readout, col, sample))
        m.add_command(label="이 Read-out Marker만 삭제",
                      command=lambda: (self.model.delete_point(group, readout, col, sample),
                                       self._redraw()))
        m.add_command(label="Sample 전체 삭제 (모든 Read-out)",
                      command=lambda: (self.model.delete_sample_all_readouts(group, col, sample),
                                       self._redraw()))
        m.tk_popup(self.winfo_pointerx(), self.winfo_pointery())

    def _change_color(self, group, readout, col, sample):
        c = colorchooser.askcolor(parent=self)[1]
        if c:
            self.model.set_color(group, readout, col, sample, c)
            self._redraw()

    def _set_ylim(self):
        group, col = self.selected[self.cur_idx]
        cur = self.model.ylim.get((group, col), (None, None))
        dlg = tk.Toplevel(self)
        dlg.title("Y축 Min/Max")
        dlg.transient(self)
        dlg.grab_set()
        ttk.Label(dlg, text=self._label((group, col)), font=("", 9, "bold")).grid(
            row=0, column=0, columnspan=2, padx=10, pady=(10, 5))
        # Y축과 동일한 배치: Max가 위, Min이 아래
        ttk.Label(dlg, text="Y Max:").grid(row=1, column=0, sticky="e", padx=5)
        vmax = tk.StringVar(value="" if cur[1] is None else str(cur[1]))
        ttk.Entry(dlg, textvariable=vmax, width=15).grid(row=1, column=1, padx=10, pady=3)
        ttk.Label(dlg, text="Y Min:").grid(row=2, column=0, sticky="e", padx=5)
        vmin = tk.StringVar(value="" if cur[0] is None else str(cur[0]))
        ttk.Entry(dlg, textvariable=vmin, width=15).grid(row=2, column=1, padx=10, pady=3)

        def apply():
            try:
                ymin = float(vmin.get())
                ymax = float(vmax.get())
            except ValueError:
                messagebox.showwarning("알림", "숫자를 입력하세요.", parent=dlg)
                return
            if ymax <= ymin:
                messagebox.showwarning("알림", "Y Max는 Y Min보다 커야 합니다.", parent=dlg)
                return
            self.model.set_ylim(group, col, ymin, ymax)
            dlg.destroy()
            self._redraw()

        def reset():
            self.model.set_ylim(group, col, None, None)
            dlg.destroy()
            self._redraw()

        btns = ttk.Frame(dlg)
        btns.grid(row=3, column=0, columnspan=2, pady=10)
        ttk.Button(btns, text="적용", command=apply).pack(side="left", padx=5)
        ttk.Button(btns, text="자동(초기화)", command=reset).pack(side="left", padx=5)
        ttk.Button(btns, text="취소", command=dlg.destroy).pack(side="left", padx=5)
        center_window(dlg)

    def _limit_dialog(self):
        """High/Low limit: 값만 입력. 이름/색(빨강 점선)은 고정."""
        group, col = self.selected[self.cur_idx]
        cur = self.model.limits.get((group, col), {})
        dlg = tk.Toplevel(self)
        dlg.title("High/Low Limit")
        dlg.transient(self)
        dlg.grab_set()
        ttk.Label(dlg, text=self._label((group, col)), font=("", 9, "bold")).grid(
            row=0, column=0, columnspan=2, padx=10, pady=(10, 5))
        ttk.Label(dlg, text="High limit:").grid(row=1, column=0, sticky="e", padx=5)
        vh = tk.StringVar(value="" if cur.get("high") is None else str(cur["high"]))
        ttk.Entry(dlg, textvariable=vh, width=14).grid(row=1, column=1, padx=10, pady=3)
        ttk.Label(dlg, text="Low limit:").grid(row=2, column=0, sticky="e", padx=5)
        vl = tk.StringVar(value="" if cur.get("low") is None else str(cur["low"]))
        ttk.Entry(dlg, textvariable=vl, width=14).grid(row=2, column=1, padx=10, pady=3)
        ttk.Label(dlg, text="※ 빈칸이면 해당 limit 없음\n"
                            "※ limit 밖 data는 * 마커로 표시됩니다",
                  foreground="gray").grid(row=3, column=0, columnspan=2, pady=3)

        def apply():
            def parse(sv):
                t = sv.get().strip()
                if not t:
                    return None
                return float(t)
            try:
                hi, lo = parse(vh), parse(vl)
            except ValueError:
                messagebox.showwarning("알림", "숫자를 입력하세요.", parent=dlg)
                return
            if hi is not None and lo is not None and hi <= lo:
                messagebox.showwarning("알림", "High limit는 Low limit보다 커야 합니다.",
                                       parent=dlg)
                return
            self.model.set_limits(group, col, hi, lo)
            dlg.destroy()
            self._redraw()

        btns = ttk.Frame(dlg)
        btns.grid(row=4, column=0, columnspan=2, pady=10)
        ttk.Button(btns, text="적용", command=apply).pack(side="left", padx=5)
        ttk.Button(btns, text="취소", command=dlg.destroy).pack(side="left", padx=5)
        center_window(dlg)

    STYLES = [("실선", "-"), ("점선", ":"), ("파선", "--"), ("일점쇄선", "-.")]

    def _custom_line_dialog(self):
        """임의의 선: X/Y축, 값, 모양, 색, 이름(선택). 수정/삭제 가능."""
        group, col = self.selected[self.cur_idx]
        kinds = [("Read-out graph", "line"), ("Delta % graph", "delta"), ("Box plot", "box")]
        try:
            cur_tab = self.nb.index(self.nb.select())
        except Exception:
            cur_tab = 0
        # 통합 탭(0)이면 기본 line, Box 탭(1)이면 box
        kind_default = 0 if cur_tab == 0 else 2

        dlg = tk.Toplevel(self)
        dlg.title("임의의 선 관리")
        dlg.transient(self)
        dlg.grab_set()
        ttk.Label(dlg, text=self._label((group, col)), font=("", 9, "bold")).grid(
            row=0, column=0, columnspan=4, padx=10, pady=(10, 5))
        ttk.Label(dlg, text="대상 그래프:").grid(row=1, column=0, sticky="e", padx=5)
        kind_cb = ttk.Combobox(dlg, state="readonly", width=16,
                               values=[k[0] for k in kinds])
        kind_cb.current(kind_default)
        kind_cb.grid(row=1, column=1, columnspan=3, sticky="w", pady=2)

        lb = tk.Listbox(dlg, height=5, width=52)
        lb.grid(row=2, column=0, columnspan=4, padx=10, pady=5)

        def cur_kind():
            return kinds[kind_cb.current()][1]

        def cur_lines():
            return list(self.model.custom_lines.get((group, col, cur_kind()), []))

        def refresh():
            lb.delete(0, "end")
            for cl in cur_lines():
                style_name = next((n for n, s in self.STYLES if s == cl.get("style")), "?")
                lb.insert("end", f"[{cl['axis'].upper()}={cl['value']}] "
                                 f"{cl.get('name','')} {style_name} ({cl.get('color')})")
        kind_cb.bind("<<ComboboxSelected>>", lambda e: refresh())

        ttk.Label(dlg, text="축:").grid(row=3, column=0, sticky="e", padx=5)
        axis_cb = ttk.Combobox(dlg, state="readonly", width=4, values=["Y", "X"])
        axis_cb.current(0)
        axis_cb.grid(row=3, column=1, sticky="w")
        ttk.Label(dlg, text="값:").grid(row=3, column=2, sticky="e", padx=5)
        val_var = tk.StringVar()
        ttk.Entry(dlg, textvariable=val_var, width=10).grid(row=3, column=3, sticky="w")
        ttk.Label(dlg, text="모양:").grid(row=4, column=0, sticky="e", padx=5)
        style_cb = ttk.Combobox(dlg, state="readonly", width=8,
                                values=[n for n, _ in self.STYLES])
        style_cb.current(1)
        style_cb.grid(row=4, column=1, sticky="w")
        ttk.Label(dlg, text="이름(선택):").grid(row=4, column=2, sticky="e", padx=5)
        name_var = tk.StringVar()
        ttk.Entry(dlg, textvariable=name_var, width=12).grid(row=4, column=3, sticky="w")
        color_var = tk.StringVar(value="#1f77b4")
        color_btn = tk.Button(dlg, text="색 선택", bg=color_var.get(), width=8,
                              command=lambda: _pick())
        color_btn.grid(row=5, column=0, columnspan=2, sticky="e", padx=5, pady=3)

        def _pick():
            c = colorchooser.askcolor(color=color_var.get(), parent=dlg)[1]
            if c:
                color_var.set(c)
                color_btn.config(bg=c)

        def make_item():
            try:
                v = float(val_var.get())
            except ValueError:
                messagebox.showwarning("알림", "값에 숫자를 입력하세요.", parent=dlg)
                return None
            return dict(axis=axis_cb.get().lower(), value=v,
                        style=self.STYLES[style_cb.current()][1],
                        color=color_var.get(), name=name_var.get().strip())

        def add():
            it = make_item()
            if it is None:
                return
            lines = cur_lines()
            lines.append(it)
            self.model.set_custom_lines(group, col, cur_kind(), lines)
            refresh()
            self._redraw()

        def load_sel(_e=None):
            sel = lb.curselection()
            if not sel:
                return
            cl = cur_lines()[sel[0]]
            axis_cb.current(0 if cl["axis"] == "y" else 1)
            val_var.set(str(cl["value"]))
            idx = next((i for i, (n, s) in enumerate(self.STYLES)
                        if s == cl.get("style")), 1)
            style_cb.current(idx)
            name_var.set(cl.get("name", ""))
            color_var.set(cl.get("color", "#1f77b4"))
            color_btn.config(bg=color_var.get())
        lb.bind("<<ListboxSelect>>", load_sel)

        def modify():
            sel = lb.curselection()
            if not sel:
                messagebox.showwarning("알림", "수정할 선을 목록에서 선택하세요.", parent=dlg)
                return
            it = make_item()
            if it is None:
                return
            lines = cur_lines()
            lines[sel[0]] = it
            self.model.set_custom_lines(group, col, cur_kind(), lines)
            refresh()
            self._redraw()

        def remove():
            sel = lb.curselection()
            if not sel:
                return
            lines = cur_lines()
            del lines[sel[0]]
            self.model.set_custom_lines(group, col, cur_kind(), lines)
            refresh()
            self._redraw()

        btns = ttk.Frame(dlg)
        btns.grid(row=6, column=0, columnspan=4, pady=10)
        ttk.Button(btns, text="추가", command=add).pack(side="left", padx=4)
        ttk.Button(btns, text="선택 수정", command=modify).pack(side="left", padx=4)
        ttk.Button(btns, text="선택 삭제", command=remove).pack(side="left", padx=4)
        ttk.Button(btns, text="닫기", command=dlg.destroy).pack(side="left", padx=4)
        refresh()
        center_window(dlg)

    def _text_dialog(self):
        """임의 텍스트: 내용/크기/색. 추가 후 그래프에서 드래그로 이동."""
        group, col = self.selected[self.cur_idx]
        kinds = [("Read-out graph", "line"), ("Delta % graph", "delta"), ("Box plot", "box")]
        try:
            cur_tab = self.nb.index(self.nb.select())
        except Exception:
            cur_tab = 0
        kind_default = 0 if cur_tab == 0 else 2

        dlg = tk.Toplevel(self)
        dlg.title("텍스트 관리")
        dlg.transient(self)
        dlg.grab_set()
        ttk.Label(dlg, text=self._label((group, col)), font=("", 9, "bold")).grid(
            row=0, column=0, columnspan=4, padx=10, pady=(10, 5))
        ttk.Label(dlg, text="대상 그래프:").grid(row=1, column=0, sticky="e", padx=5)
        kind_cb = ttk.Combobox(dlg, state="readonly", width=16,
                               values=[k[0] for k in kinds])
        kind_cb.current(kind_default)
        kind_cb.grid(row=1, column=1, columnspan=3, sticky="w", pady=2)

        lb = tk.Listbox(dlg, height=5, width=52)
        lb.grid(row=2, column=0, columnspan=4, padx=10, pady=5)

        def cur_kind():
            return kinds[kind_cb.current()][1]

        def cur_texts():
            return list(self.model.texts.get((group, col, cur_kind()), []))

        def refresh():
            lb.delete(0, "end")
            for t in cur_texts():
                lb.insert("end", f"'{t['text']}'  size={t.get('size',10)} "
                                 f"({t.get('color','black')})")
        kind_cb.bind("<<ComboboxSelected>>", lambda e: refresh())

        ttk.Label(dlg, text="내용:").grid(row=3, column=0, sticky="e", padx=5)
        txt_var = tk.StringVar()
        ttk.Entry(dlg, textvariable=txt_var, width=28).grid(
            row=3, column=1, columnspan=3, sticky="w", pady=3)
        ttk.Label(dlg, text="크기:").grid(row=4, column=0, sticky="e", padx=5)
        size_var = tk.StringVar(value="10")
        ttk.Entry(dlg, textvariable=size_var, width=6).grid(row=4, column=1, sticky="w")
        color_var = tk.StringVar(value="#000000")
        color_btn = tk.Button(dlg, text="색 선택", bg=color_var.get(), width=8,
                              command=lambda: _pick())
        color_btn.grid(row=4, column=2, columnspan=2, sticky="w", padx=5)

        def _pick():
            c = colorchooser.askcolor(color=color_var.get(), parent=dlg)[1]
            if c:
                color_var.set(c)
                color_btn.config(bg=c)

        def make_item(base=None):
            try:
                size = float(size_var.get())
            except ValueError:
                messagebox.showwarning("알림", "크기에 숫자를 입력하세요.", parent=dlg)
                return None
            if not txt_var.get().strip():
                messagebox.showwarning("알림", "내용을 입력하세요.", parent=dlg)
                return None
            it = dict(base or {})
            it.update(text=txt_var.get(), size=size, color=color_var.get())
            return it

        def add():
            # 초기 위치: 그래프 중앙 (이후 드래그로 이동)
            ax = self._ax_line if cur_kind() == "line" else (
                self._ax_delta if cur_kind() == "delta" else None)
            if ax is not None:
                x0, x1 = ax.get_xlim()
                y0, y1 = ax.get_ylim()
                cx, cy = (x0 + x1) / 2, (y0 + y1) / 2
            else:
                cx, cy = 0.5, 0.5
            it = make_item({"x": cx, "y": cy})
            if it is None:
                return
            texts = cur_texts()
            texts.append(it)
            self.model.set_texts(group, col, cur_kind(), texts)
            refresh()
            self._redraw()

        def load_sel(_e=None):
            sel = lb.curselection()
            if not sel:
                return
            t = cur_texts()[sel[0]]
            txt_var.set(t["text"])
            size_var.set(str(t.get("size", 10)))
            color_var.set(t.get("color", "#000000"))
            color_btn.config(bg=color_var.get())
        lb.bind("<<ListboxSelect>>", load_sel)

        def modify():
            sel = lb.curselection()
            if not sel:
                messagebox.showwarning("알림", "수정할 텍스트를 선택하세요.", parent=dlg)
                return
            texts = cur_texts()
            it = make_item(texts[sel[0]])
            if it is None:
                return
            texts[sel[0]] = it
            self.model.set_texts(group, col, cur_kind(), texts)
            refresh()
            self._redraw()

        def remove():
            sel = lb.curselection()
            if not sel:
                return
            texts = cur_texts()
            del texts[sel[0]]
            self.model.set_texts(group, col, cur_kind(), texts)
            refresh()
            self._redraw()

        btns = ttk.Frame(dlg)
        btns.grid(row=5, column=0, columnspan=4, pady=10)
        ttk.Button(btns, text="추가", command=add).pack(side="left", padx=4)
        ttk.Button(btns, text="선택 수정", command=modify).pack(side="left", padx=4)
        ttk.Button(btns, text="선택 삭제", command=remove).pack(side="left", padx=4)
        ttk.Button(btns, text="닫기", command=dlg.destroy).pack(side="left", padx=4)
        ttk.Label(dlg, text="※ 추가 후 그래프에서 텍스트를 드래그해 위치를 옮길 수 있습니다",
                  foreground="gray").grid(row=6, column=0, columnspan=4, pady=(0, 8))
        refresh()
        center_window(dlg)

    def _undo(self):
        if self.model.undo():
            self._redraw()

    def _redo(self):
        if self.model.redo():
            self._redraw()

    # ---- PDF ------------------------------------------------------------------
    def _export_pdf(self):
        if len(self.model.groups) == 1:
            default_name = f"{self.model.reliability}_{self.model.groups[0]}_Data_Analysis.pdf"
        else:
            default_name = f"{self.model.reliability}_Data_Analysis.pdf"
        path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            initialfile=default_name,
            filetypes=[("PDF", "*.pdf")],
            confirmoverwrite=False)  # 중복 확인은 아래 팝업에서 직접 수행
        if not path:
            return
        if os.path.exists(path):
            if not messagebox.askyesno(
                    "덮어쓰기 확인",
                    f"동일한 이름의 파일이 이미 있습니다.\n\n{os.path.basename(path)}\n\n덮어쓸까요?",
                    parent=self):
                return
        win = tk.Toplevel(self)
        win.title("PDF 생성 중")
        win.geometry("360x90")
        ttk.Label(win, text="PDF Report 생성 중...").pack(pady=8)
        pbar = ttk.Progressbar(win, mode="determinate", length=320)
        pbar.pack(pady=5)
        center_window(win, 360, 90)

        def cb(done, total):
            pbar["maximum"] = total
            pbar["value"] = done
            win.update_idletasks()

        def work():
            try:
                export_pdf(self.model, self.selected, path, cb)
                self.after(0, lambda: (win.destroy(),
                                       messagebox.showinfo("완료", f"PDF 저장 완료:\n{path}",
                                                           parent=self)))
            except Exception as e:
                err = traceback.format_exc()
                self.after(0, lambda: (win.destroy(),
                                       messagebox.showerror("오류", f"PDF 생성 실패:\n{e}\n\n{err}",
                                                            parent=self)))

        threading.Thread(target=work, daemon=True).start()


# ============================================================================
if __name__ == "__main__":
    app = App()
    app.mainloop()
